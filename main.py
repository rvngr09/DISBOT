import discord
from discord.ext import commands
import openpyxl
import logging
import os
import json
import asyncio
from datetime import datetime

# === CONFIG ===
TOKEN = os.getenv("DISCORD_TOKEN") or "MTQyMjU4Mzg0ODUzNzE2NTg2NA.G9wBli.kk3hBHRsnzx5q7MkZnwfA-Du42jYMJxoAmFBp0"  # Utiliser une variable d'environnement
if not TOKEN:
    logging.error("Token Discord non trouvé. Définir la variable DISCORD_TOKEN")
    exit(1)

GUILD_ID = 1451327990628614298
ROLE_NAME = "ACAD B"
EXCEL_FILE = "CMS62026.xlsx"
CHANNEL_ID = 1451336152568037456
CLAIM_FILE = "claimed.json"
LOG_FILE = "bot_activity.log"

# === LOGGING AVANCÉ ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# === DISCORD INTENTS ===
intents = discord.Intents.default()
intents.members = True
intents.message_content = True
intents.guilds = True

bot = commands.Bot(
    command_prefix="!",
    intents=intents,
    help_command=None  # Personnaliser l'aide
)

# === GLOBALS ===
matricules = set()
claimed = {}
excel_headers = []


# === LOAD MATRICULES WITH ERROR HANDLING ===
def load_matricules():
    """Version CORRIGÉE avec la bonne colonne de section"""
    global excel_headers

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active

        excel_headers = [cell.value for cell in sheet[1]]
        logging.info(f"En-têtes: {excel_headers}")

        # === TROUVER LES BONNES COLONNES ===
        col_indices = {}

        # 1. Colonne Matricule (colonne G, index 6)
        col_indices['matricule'] = 6  # Fixe car on sait que c'est colonne 7 (G)

        # 2. Colonne Programme (Affectation) - colonne I, index 8
        col_indices['program'] = 8  # Colonne I

        # 3. Colonne Section IMPORTANTE : "Section Prog. Web" (colonne J, index 9)
        # Celle-ci contient probablement "B", pas la colonne "Sect"
        col_indices['section'] = 9  # Colonne J ("Section Prog. Web")

        logging.info(f"Colonnes utilisées: {col_indices}")

        valid_matricules = set()
        invalid_reasons = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # === LIRE LE MATRICULE ===
                matricule_raw = row[col_indices['matricule']]
                if matricule_raw is None:
                    continue

                # Convertir proprement
                if isinstance(matricule_raw, (int, float)):
                    # Si c'est un float (212231455913.0), convertir en int puis string
                    matricule = str(int(matricule_raw))
                else:
                    matricule = str(matricule_raw).strip()
                    # Nettoyer
                    matricule = ''.join(filter(str.isdigit, matricule))

                if not matricule:
                    continue

                # === LIRE LE PROGRAMME ===
                program_raw = row[col_indices['program']]
                program = str(program_raw or "").strip().lower() if program_raw else ""

                # === LIRE LA SECTION IMPORTANTE ===
                section_raw = row[col_indices['section']]
                section = str(section_raw or "").strip().upper() if section_raw else ""

                # === DEBUG pour le matricule problématique ===
                if matricule == "212231455913":
                    logging.info(f"DEBUG {matricule}: Program='{program[:50]}...', Section='{section}'")

                # === VÉRIFICATIONS ===
                is_valid = True
                reason = ""

                # 1. Vérifier programme
                if not program:
                    is_valid = False
                    reason = "Programme vide"
                elif not ("programmation web" in program and "introduction à l'ia" in program):
                    is_valid = False
                    reason = f"Programme incorrect: {program[:50]}..."

                # 2. Vérifier section (doit être "B")
                if section != "B":
                    is_valid = False
                    reason = f"Section incorrecte: '{section}' (attendu 'B')"

                # 3. Ajouter si valide
                if is_valid:
                    valid_matricules.add(matricule)
                    if matricule == "212231455913":
                        logging.info(f"✓ MATRICULE ACCEPTÉ: {matricule}")
                else:
                    if matricule == "212231455913":
                        logging.warning(f"✗ MATRICULE REJETÉ: {matricule} - {reason}")
                    invalid_reasons.append(f"{matricule}: {reason}")

            except Exception as e:
                logging.warning(f"Erreur ligne {row_idx}: {e}")
                continue

        logging.info(f"✅ {len(valid_matricules)} matricules valides chargés")

        # Afficher quelques raisons de rejet
        if invalid_reasons:
            logging.info(f"Exemples de rejets: {invalid_reasons[:5]}")

        return valid_matricules

    except Exception as e:
        logging.error(f"❌ Erreur chargement Excel: {e}")
        return set()

    except Exception as e:
        logging.error(f"❌ Erreur chargement Excel: {e}")
        return set()


# === LOAD CLAIMED MATRICULES ===
def load_claims():
    """Charge les matricules déjà attribués"""
    if os.path.exists(CLAIM_FILE):
        try:
            with open(CLAIM_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Erreur lors du chargement des claims: {e}")
            return {}
    return {}


def save_claims():
    """Sauvegarde les matricules attribués"""
    try:
        with open(CLAIM_FILE, "w", encoding="utf-8") as f:
            json.dump(claimed, f, indent=4)
    except Exception as e:
        logging.error(f"Erreur lors de la sauvegarde des claims: {e}")


# === EVENTS ===
@bot.event
async def on_ready():
    """Événement déclenché quand le bot est prêt"""
    global matricules, claimed

    logging.info(f"Bot connecté: {bot.user.name} (ID: {bot.user.id})")
    logging.info(f"Servers: {len(bot.guilds)}")

    # Charger les données
    matricules = load_matricules()
    claimed = load_claims()

    # Vérifier la présence des fichiers nécessaires
    if not matricules:
        logging.warning("Aucun matricule chargé. Vérifiez le fichier Excel.")

    # Changer le statut du bot
    await bot.change_presence(
        activity=discord.Activity(
            type=discord.ActivityType.watching,
            name=f"{len(matricules)} matricules"
        )
    )


@bot.event
async def on_command_error(ctx, error):
    """Gestion des erreurs de commandes"""
    if isinstance(error, commands.MissingPermissions):
        await ctx.send("❌ Vous n'avez pas la permission d'utiliser cette commande.")
    elif isinstance(error, commands.MissingRequiredArgument):
        await ctx.send(f"❌ Argument manquant. Usage: `{ctx.prefix}{ctx.command.name} {ctx.command.signature}`")
    else:
        logging.error(f"Erreur de commande: {error}")
        await ctx.send("❌ Une erreur est survenue.")


# === MAIN VALIDATION LOGIC ===
@bot.event
async def on_message(message):
    """Validation des matricules envoyés dans le channel dédié"""
    if message.author.bot:
        return

    # Vérifier si c'est le bon channel
    if message.channel.id != CHANNEL_ID:
        await bot.process_commands(message)
        return

    user_input = message.content.strip().upper()

    # Nettoyer l'input
    matricule = ''.join(c for c in user_input if c.isalnum())

    if not matricule:
        await message.channel.send(f"{message.author.mention}, veuillez entrer un matricule valide.")
        return

    logging.info(f"Validation tentative: {message.author} -> '{matricule}'")

    try:
        # Récupérer le rôle
        guild = message.guild
        role = discord.utils.get(guild.roles, name=ROLE_NAME)

        if not role:
            logging.error(f"Rôle '{ROLE_NAME}' introuvable")
            await message.channel.send("❌ Erreur: rôle non configuré.")
            return

        if matricule in matricules:
            # Vérifier si déjà attribué
            if matricule in claimed:
                claimant_id = claimed[matricule]

                if claimant_id == str(message.author.id):
                    # Même utilisateur
                    if role not in message.author.roles:
                        await message.author.add_roles(role)
                        await message.channel.send(
                            f"{message.author.mention}, matricule déjà validé ✅. Rôle {ROLE_NAME} ajouté."
                        )
                    else:
                        await message.channel.send(
                            f"{message.author.mention}, tu as déjà validé ton matricule ✅."
                        )
                else:
                    # Tentative de fraude
                    logging.warning(f"Tentative de fraude: {message.author} tente d'utiliser le matricule {matricule}")
                    await message.channel.send(
                        f"{message.author.mention}, ce matricule est déjà utilisé par un autre membre ❌.\n"
                        f"Contactez un administrateur si c'est une erreur."
                    )
                    return

            else:
                # Nouvelle validation
                claimed[matricule] = str(message.author.id)
                save_claims()

                await message.author.add_roles(role)
                await message.channel.send(
                    f"{message.author.mention}, matricule valide ✅ ! Rôle {ROLE_NAME} attribué."
                )
                logging.info(f"Matricule {matricule} attribué à {message.author}")

        else:
            # Matricule invalide
            if role in message.author.roles:
                await message.author.remove_roles(role)
                await message.channel.send(
                    f"{message.author.mention}, matricule invalide ❌. Rôle {ROLE_NAME} retiré."
                )
                logging.info(f"Rôle retiré pour {message.author} (matricule invalide)")
            else:
                await message.channel.send(
                    f"{message.author.mention}, matricule non reconnu ❌.\n"
                    f"Vérifiez votre matricule ou contactez un enseignant."
                )

    except discord.Forbidden:
        logging.error("Permissions insuffisantes pour gérer les rôles")
        await message.channel.send("❌ Erreur de permissions. Vérifiez les droits du bot.")
    except Exception as e:
        logging.error(f"Erreur lors de la validation: {e}")
        await message.channel.send("❌ Une erreur est survenue lors de la validation.")

    await bot.process_commands(message)


# === ADMIN COMMANDS ===
@bot.command(name="checkcolumns")
@commands.has_permissions(administrator=True)
async def check_columns(ctx):
    """Vérifie les valeurs dans les colonnes de section"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active

        # Trouver toutes les colonnes liées à "section"
        section_columns = []
        for idx, header in enumerate(sheet[1], start=1):
            if header.value and any(word in str(header.value).lower() for word in ['section', 'sect']):
                section_columns.append((idx, header.value))

        # Pour le matricule 212231455913 (ligne 320)
        target_row = 320
        data = []

        for col_idx, header in section_columns:
            value = sheet.cell(row=target_row, column=col_idx).value
            data.append(f"**{header}** (colonne {col_idx}): `{value}`")

        # Aussi vérifier la colonne "Section Prog. Web"
        for idx, header in enumerate(sheet[1], start=1):
            if header.value and 'prog. web' in str(header.value).lower():
                value = sheet.cell(row=target_row, column=idx).value
                data.append(f"**{header.value}** (colonne {idx}): `{value}`")

        embed = discord.Embed(
            title="🔍 Analyse des colonnes Section",
            description=f"Pour le matricule 212231455913 (ligne {target_row}):",
            color=discord.Color.orange()
        )

        embed.add_field(
            name="📊 Valeurs trouvées",
            value="\n".join(data),
            inline=False
        )

        # Vérifier aussi quelques autres lignes
        sample_data = []
        for row in range(2, 7):  # Lignes 2 à 6
            matricule = sheet.cell(row=row, column=7).value
            sect = sheet.cell(row=row, column=8).value if len(section_columns) > 0 else "N/A"
            section_prog = None

            # Trouver "Section Prog. Web"
            for idx, header in enumerate(sheet[1], start=1):
                if header.value and 'prog. web' in str(header.value).lower():
                    section_prog = sheet.cell(row=row, column=idx).value
                    break

            sample_data.append(f"L{row}: Mat=`{matricule}`, Sect=`{sect}`, Prog.Web=`{section_prog}`")

        embed.add_field(
            name="📝 Exemple autres lignes",
            value="\n".join(sample_data),
            inline=False
        )

        await ctx.send(embed=embed)

    except Exception as e:
        await ctx.send(f"❌ Erreur: {e}")


@bot.command(name="find")
@commands.has_permissions(administrator=True)
async def find_matricule(ctx, matricule: str):
    """Recherche un matricule dans TOUTES les cellules"""
    matricule = matricule.strip()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active

        matches = []

        # Parcourir TOUTES les cellules
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx).value
                if cell and matricule in str(cell):
                    # Trouvé ! Récupérer toute la ligne
                    row_data = []
                    for c in range(1, sheet.max_column + 1):
                        header = sheet.cell(row=1, column=c).value or f"Col{c}"
                        value = sheet.cell(row=row_idx, column=c).value
                        row_data.append(f"**{header}:** `{value}`")

                    matches.append({
                        'row': row_idx,
                        'col': col_idx,
                        'data': row_data
                    })

        if matches:
            embed = discord.Embed(
                title=f"🔍 Matricule trouvé: {matricule}",
                description=f"**{len(matches)} occurrence(s) trouvée(s)**",
                color=discord.Color.green()
            )

            for i, match in enumerate(matches[:3]):  # Limiter à 3 résultats
                embed.add_field(
                    name=f"📍 Ligne {match['row']}, Colonne {match['col']}",
                    value="\n".join(match['data'][:8]),  # Limiter à 8 champs
                    inline=False
                )

            if len(matches) > 3:
                embed.set_footer(text=f"... et {len(matches) - 3} autres occurrences")

        else:
            embed = discord.Embed(
                title=f"❌ Matricule NON trouvé: {matricule}",
                description="Le matricule n'existe nulle part dans le fichier Excel",
                color=discord.Color.red()
            )

        await ctx.send(embed=embed)

    except Exception as e:
        await ctx.send(f"❌ Erreur: {e}")


@bot.command(name="checkall")
@commands.has_permissions(administrator=True)
async def check_all_matricules(ctx):
    """Vérifie tous les matricules et montre lesquels sont valides"""
    valid_count = 0
    invalid_details = []

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active

        # Trouver les indices des colonnes
        col_indices = {}
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for i, header in enumerate(headers):
            header_str = str(header).strip() if header else ""
            if "affectation" in header_str.lower():
                col_indices["affectation"] = i
            if "section prog. web" in header_str.lower():
                col_indices["section"] = i
            if "matricule" in header_str.lower():
                col_indices["matricule"] = i

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                matricule = str(row[col_indices["matricule"]] or "").strip().upper()
                matricule = ''.join(c for c in matricule if c.isalnum())

                if not matricule:
                    continue

                program = str(row[col_indices["affectation"]] or "").strip()
                section = str(row[col_indices["section"]] or "").strip()

                program_lower = program.lower()
                is_valid = ("programmation web" in program_lower and
                            "introduction à l'ia" in program_lower and
                            section.upper() == "B")

                if is_valid:
                    valid_count += 1
                else:
                    invalid_details.append(
                        f"Ligne {row_idx}: `{matricule}` - Prog: `{program[:30]}...` - Sect: `{section}`")

            except Exception as e:
                invalid_details.append(f"Ligne {row_idx}: ERREUR - {e}")
                continue

        # Créer l'embed de rapport
        embed = discord.Embed(
            title="📊 Rapport de Validation des Matricules",
            color=discord.Color.blue()
        )

        embed.add_field(name="✅ Matricules Valides", value=str(valid_count), inline=True)
        embed.add_field(name="❌ Matricules Invalides", value=str(len(invalid_details)), inline=True)
        embed.add_field(name="📈 Total", value=str(valid_count + len(invalid_details)), inline=True)

        if invalid_details:
            # Limiter à 10 lignes pour ne pas dépasser la limite Discord
            details_text = "\n".join(invalid_details[:10])
            if len(invalid_details) > 10:
                details_text += f"\n... et {len(invalid_details) - 10} autres"

            embed.add_field(
                name="📝 Détails des Invalides",
                value=f"```{details_text}```",
                inline=False
            )

        await ctx.send(embed=embed)

    except Exception as e:
        logging.error(f"Erreur checkall: {e}")
        await ctx.send(f"❌ Erreur: {e}")
# === ERROR HANDLER ===
@bot.event
async def on_error(event, *args, **kwargs):
    logging.error(f"Erreur dans l'événement {event}: {args} {kwargs}")


# === RUN BOT ===
if __name__ == "__main__":
    try:
        bot.run(TOKEN)
    except discord.LoginFailure:
        logging.error("Échec de connexion. Vérifiez le token Discord.")
    except Exception as e:
        logging.error(f"Erreur inattendue: {e}")